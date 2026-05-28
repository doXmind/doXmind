"""Direct backend benchmark driver.

Skips the FastAPI / Tauri / browser layers and calls the parse/read services
in-process. Forces DOXMIND_PERF=1 + a sandbox log path so it never pollutes
the user's real ~/.doxmind/perf.log.

Usage:
    cd server && .venv/bin/python ../scripts/perf-bench.py

What it does:
  1. Ensures fixtures exist (small.md, large.md, huge.md, synthetic.pdf,
     synthetic.xlsx) under testdata/perf/. Generates anything missing.
  2. Runs each scenario N=5 times, throws away the first (cache/JIT warm-up)
     and reports median + p95 in human-readable form.
  3. Aggregates the JSON-lines log via scripts/perf-summary.mjs at the end.
"""

from __future__ import annotations

import json
import os
import statistics
import subprocess
import sys
import time
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
SERVER_DIR = REPO_ROOT / "server"
FIXTURES = REPO_ROOT / "testdata" / "perf"
PERF_LOG = FIXTURES / "_bench.log"

# Configure perf BEFORE importing any service modules so the timing helper
# picks up the env var at import time.
os.environ["DOXMIND_PERF"] = "1"
os.environ["DOXMIND_PERF_LOG"] = str(PERF_LOG)

# Make sure the server package is importable.
sys.path.insert(0, str(SERVER_DIR))

from services.excel_workbook import (  # noqa: E402
    _clear_xlsx_cache,
    parse_workbook_json_bytes,
)
from services.markdown_document_state import (  # noqa: E402
    MarkdownDocumentState,
    _clear_read_cache,
)
from services.pdf_blocks import _clear_pdf_cache, parse_pdf_blocks  # noqa: E402

WARMUP = 1
RUNS = 5


def clear_all_caches() -> None:
    _clear_read_cache()
    _clear_pdf_cache()
    _clear_xlsx_cache()


def _print(line: str) -> None:
    sys.stdout.write(line + "\n")
    sys.stdout.flush()


# ---------------------------------------------------------------- fixtures


def ensure_md_fixtures() -> dict[str, Path]:
    """Generate small / large / huge markdown fixtures if missing."""
    FIXTURES.mkdir(parents=True, exist_ok=True)
    sizes = {
        "small": (100, 0, 0, 0, 0),  # lines, math, mermaid, db, links
        "large": (6000, 60, 5, 3, 200),
        "huge": (30000, 300, 20, 10, 1000),
    }
    paths: dict[str, Path] = {}
    for name, (lines, math, mermaid, db, links) in sizes.items():
        out = FIXTURES / f"{name}.md"
        if not out.exists() or out.stat().st_size < lines * 30:
            out.write_text(_build_md(name, lines, math, mermaid, db, links))
        paths[name] = out
    return paths


def _build_md(name: str, lines: int, math: int, mermaid: int, db: int, links: int) -> str:
    out = [
        "---",
        f"title: Perf {name}",
        f"id: {hex(abs(hash(name)) & 0xFFFFFFFFFFFFFFFF)[2:]}-{name}",
        "---",
        "",
        f"# Perf {name}",
        "",
    ]

    def para(idx: int) -> str:
        seed = (idx * 2654435761) & 0xFFFFFFFF
        rotor = "abcdefghijklmnopqrstuvwxyz"
        words = []
        for i in range(20):
            wlen = 3 + ((seed >> (i % 11)) & 7)
            w = "".join(rotor[(seed + j * (i + 1)) % len(rotor)] for j in range(wlen))
            words.append(w)
        return " ".join(words) + "."

    def every(total: int) -> int:
        return max(1, lines // max(total, 1))

    me, mr, dbe, lk = every(math), every(mermaid), every(db), every(links)
    mi = mri = dbi = lki = 0
    for i in range(lines):
        if mi < math and i and i % me == 0:
            out += ["", "$$", f"f_{{{mi}}}(x) = \\sum_{{k=0}}^{{{mi+3}}} \\frac{{x^k}}{{k!}}", "$$", ""]
            mi += 1
            continue
        if mri < mermaid and i and i % mr == 0:
            out += ["", "```mermaid", "flowchart LR", f"  A{mri} --> B{mri} --> C{mri}", "```", ""]
            mri += 1
            continue
        if dbi < db and i and i % dbe == 0:
            id_ = f"db{dbi:03d}-0000-0000-0000-000000000000"
            out += ["", f"<!-- database:{id_} -->", f'<div data-database-id="{id_}"></div>', ""]
            dbi += 1
            continue
        if lki < links and i and i % lk == 0:
            out.append(f"See also: [[link-target-{lki}]]")
            lki += 1
            continue
        if i % 80 == 0 and i:
            out += ["", f"## Section {i // 80}", ""]
        out.append(para(i))
    return "\n".join(out)


def ensure_pdf(name: str, pages: int) -> Path:
    out = FIXTURES / f"{name}.pdf"
    if out.exists():
        return out
    import pymupdf

    doc = pymupdf.open()
    for i in range(pages):
        page = doc.new_page(width=595, height=842)  # A4 in pt
        # Sprinkle a few paragraphs per page so PyMuPDF has to cluster
        # blocks/lines/spans non-trivially.
        for j in range(8):
            page.insert_text(
                (50, 60 + j * 80),
                f"Page {i+1} block {j}: " + " ".join(f"word{k}" for k in range(40)),
                fontsize=10,
            )
    doc.save(str(out))
    doc.close()
    return out


def ensure_xlsx(name: str, sheets: int, rows: int, cols: int) -> Path:
    out = FIXTURES / f"{name}.xlsx"
    if out.exists():
        return out
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    for s in range(sheets):
        ws = wb.create_sheet(f"Sheet{s+1}")
        for r in range(rows):
            for c in range(cols):
                if c == 0:
                    ws.cell(row=r + 1, column=c + 1, value=f"R{r}")
                elif c == cols - 1:
                    # Last column is a formula referencing the row sum
                    ws.cell(row=r + 1, column=c + 1, value=f"=SUM(B{r+1}:{chr(ord('A')+cols-2)}{r+1})")
                else:
                    ws.cell(row=r + 1, column=c + 1, value=(r * cols + c) % 997)
    wb.save(str(out))
    wb.close()
    return out


def ensure_xlsx_no_formulas(name: str, sheets: int, rows: int, cols: int) -> Path:
    out = FIXTURES / f"{name}.xlsx"
    if out.exists():
        return out
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    for s in range(sheets):
        ws = wb.create_sheet(f"Sheet{s+1}")
        for r in range(rows):
            for c in range(cols):
                ws.cell(row=r + 1, column=c + 1, value=(r * cols + c) % 997)
    wb.save(str(out))
    wb.close()
    return out


# ---------------------------------------------------------------- runners


def time_n(
    fn,
    *,
    label: str,
    n: int = RUNS,
    warmup: int = WARMUP,
    before_each=None,
) -> dict[str, float]:
    for _ in range(warmup):
        if before_each is not None:
            before_each()
        fn()
    samples = []
    for _ in range(n):
        if before_each is not None:
            before_each()
        t0 = time.perf_counter()
        fn()
        samples.append((time.perf_counter() - t0) * 1000)
    samples.sort()
    return {
        "label": label,
        "n": n,
        "min": samples[0],
        "median": statistics.median(samples),
        "p95": samples[int(0.95 * (n - 1))],
        "max": samples[-1],
    }


def main() -> int:
    if PERF_LOG.exists():
        PERF_LOG.unlink()
    md_paths = ensure_md_fixtures()
    medium_pdf = ensure_pdf("medium", 30)
    large_pdf = ensure_pdf("large", 200)
    medium_xlsx = ensure_xlsx("medium", sheets=10, rows=2000, cols=20)
    medium_flat_xlsx = ensure_xlsx_no_formulas("medium-flat", sheets=5, rows=1000, cols=20)
    large_xlsx = ensure_xlsx("large", sheets=10, rows=5000, cols=50)
    state = MarkdownDocumentState()

    results: list[dict] = []

    _print("=== fixtures ===")
    for k, p in md_paths.items():
        _print(f"  {k:<6} {p.stat().st_size / 1024:>8.1f} KB  {p}")
    _print(f"  pdf-medium {medium_pdf.stat().st_size / 1024:>8.1f} KB  {medium_pdf}")
    _print(f"  pdf-large  {large_pdf.stat().st_size / 1024:>8.1f} KB  {large_pdf}")
    _print(f"  xlsx-medium {medium_xlsx.stat().st_size / 1024:>8.1f} KB  {medium_xlsx}")
    _print(
        f"  xlsx-flat   {medium_flat_xlsx.stat().st_size / 1024:>8.1f} KB  {medium_flat_xlsx}"
    )
    _print(f"  xlsx-large  {large_xlsx.stat().st_size / 1024:>8.1f} KB  {large_xlsx}")
    _print("")

    # Markdown: three scenarios per fixture
    #   * cold-no-cache: clear cache + delete sidecar before each call. This
    #     is the "first time the user ever opens this never-saved file"
    #     baseline. The optimization should NOT regress this.
    #   * repeat-cache-hit: do not clear cache. Models the "user just opened
    #     this same file again" path. The optimization should make this
    #     near-zero.
    for tag, path in md_paths.items():
        sidecar = path.parent / f".{path.name.replace('.md', '.doxmind')}"

        def cold_no_cache(p=path, s=sidecar):
            if s.exists():
                s.unlink()
            p.write_text(p.read_text())  # bump mtime to defeat cache key
            clear_all_caches()
            state.read(p)

        results.append(time_n(cold_no_cache, label=f"md.{tag}.cold-no-cache"))

        def repeat_hit(p=path):
            state.read(p)

        # Prime cache once before the timed runs so all 5 hits are warm.
        state.read(path)
        results.append(time_n(repeat_hit, label=f"md.{tag}.repeat-cache-hit"))

    # PDF: same pattern. Bytes are loaded once outside the runner.
    medium_bytes = medium_pdf.read_bytes()
    large_bytes = large_pdf.read_bytes()

    def pdf_cold_medium(b=medium_bytes):
        clear_all_caches()
        parse_pdf_blocks(b)

    def pdf_cold_large(b=large_bytes):
        clear_all_caches()
        parse_pdf_blocks(b)

    results.append(time_n(pdf_cold_medium, label="pdf.medium.cold-no-cache"))
    results.append(time_n(pdf_cold_large, label="pdf.large.cold-no-cache"))
    parse_pdf_blocks(medium_bytes)
    parse_pdf_blocks(large_bytes)
    results.append(
        time_n(lambda b=medium_bytes: parse_pdf_blocks(b), label="pdf.medium.repeat-cache-hit")
    )
    results.append(
        time_n(lambda b=large_bytes: parse_pdf_blocks(b), label="pdf.large.repeat-cache-hit")
    )

    # Excel
    medium_xlsx_bytes = medium_xlsx.read_bytes()
    medium_flat_xlsx_bytes = medium_flat_xlsx.read_bytes()
    large_xlsx_bytes = large_xlsx.read_bytes()

    def excel_cold_medium(b=medium_xlsx_bytes):
        clear_all_caches()
        parse_workbook_json_bytes(b)

    def excel_cold_medium_flat(b=medium_flat_xlsx_bytes):
        clear_all_caches()
        parse_workbook_json_bytes(b)

    def excel_cold_large(b=large_xlsx_bytes):
        clear_all_caches()
        parse_workbook_json_bytes(b)

    results.append(time_n(excel_cold_medium, label="excel.medium.cold-no-cache", n=3))
    results.append(time_n(excel_cold_medium_flat, label="excel.medium-flat.cold-no-cache", n=3))
    results.append(time_n(excel_cold_large, label="excel.large.cold-no-cache", n=3))
    parse_workbook_json_bytes(medium_xlsx_bytes)
    parse_workbook_json_bytes(medium_flat_xlsx_bytes)
    parse_workbook_json_bytes(large_xlsx_bytes)
    results.append(
        time_n(
            lambda b=medium_xlsx_bytes: parse_workbook_json_bytes(b),
            label="excel.medium.repeat-cache-hit",
        )
    )
    results.append(
        time_n(
            lambda b=medium_flat_xlsx_bytes: parse_workbook_json_bytes(b),
            label="excel.medium-flat.repeat-cache-hit",
        )
    )
    results.append(
        time_n(
            lambda b=large_xlsx_bytes: parse_workbook_json_bytes(b),
            label="excel.large.repeat-cache-hit",
        )
    )

    # ---- print table ----
    _print("=== top-level wall clock (ms) ===")
    _print(f"{'scenario':<28} {'n':>3} {'min':>9} {'med':>9} {'p95':>9} {'max':>9}")
    _print("-" * 72)
    for r in results:
        _print(
            f"{r['label']:<28} {r['n']:>3} "
            f"{r['min']:>9.1f} {r['median']:>9.1f} "
            f"{r['p95']:>9.1f} {r['max']:>9.1f}"
        )
    _print("")

    # ---- aggregate finer-grained span log ----
    _print("=== aggregated span log (ms; from DOXMIND_PERF backend log) ===")
    summary = subprocess.run(
        ["node", str(REPO_ROOT / "scripts" / "perf-summary.mjs"), str(PERF_LOG)],
        capture_output=True,
        text=True,
        check=False,
    )
    if summary.returncode == 0:
        sys.stdout.write(summary.stdout)
    else:
        _print("(perf-summary failed)")
        _print(summary.stderr)

    # Also dump JSON for further analysis
    out_json = FIXTURES / "_bench-results.json"
    out_json.write_text(json.dumps(results, indent=2))
    _print("")
    _print(f"wrote raw results to {out_json}")
    _print(f"wrote span log to {PERF_LOG}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
