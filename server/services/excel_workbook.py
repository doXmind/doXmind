"""Read-only Excel conversion backed by openpyxl.

The CLI and MCP conversion surfaces use this module to parse Excel attachments
into a bounded, JSON-serialisable cell model. It does not write workbooks or
persist editor state.
"""

from __future__ import annotations

import csv
import hashlib
import io
import logging
import os
import xml.etree.ElementTree as ET
from collections import OrderedDict
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import PurePosixPath
from threading import Lock
from typing import Any

# Cache encode/decode: prefer orjson (Rust-backed, ~2x stdlib on multi-MB
# DTOs). Falls back to stdlib json if orjson is somehow unavailable so the
# server still works even with a stripped-down env.
try:
    import orjson as _json  # type: ignore[import-not-found]

    def _cache_encode(obj: Any) -> bytes:
        return _json.dumps(obj)

    def _cache_decode(data: bytes) -> Any:
        return _json.loads(data)

except ImportError:  # pragma: no cover — orjson is in requirements.txt
    import json as _stdlib_json

    def _cache_encode(obj: Any) -> bytes:
        return _stdlib_json.dumps(obj).encode("utf-8")

    def _cache_decode(data: bytes) -> Any:
        return _stdlib_json.loads(data)


from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles.numbers import is_date_format
from openpyxl.utils import range_boundaries
from openpyxl.utils.datetime import from_excel
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from lib.timing import record as perf_record
from lib.timing import timed as perf_timed

logger = logging.getLogger(__name__)

# Defensive caps. A pathological workbook can exceed Excel's nominal
# 1,048,576 rows × 16,384 columns; even normal sheets often have used_range
# stretching off into empty territory. We clamp to a reasonable viewport so
# the JSON payload stays bounded.
MAX_SHEETS = 64
MAX_ROWS_PER_SHEET = 5000
MAX_COLS_PER_SHEET = 200
_SHEET_LAYOUT_MARKERS = (
    b"<cols",
    b":cols",
    b"<mergeCell",
    b":mergeCell",
    b"<pane",
    b":pane",
    b"<hyperlink",
    b":hyperlink",
    b" ht=",
)
_REL_ID_ATTR = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"

# Process-local LRU cache for parsed workbook DTOs. The dual openpyxl load
# in parse_workbook is expensive (15+ seconds on an 8 MB workbook); repeated
# CLI/MCP conversions in one process should not pay it again. Hashing the input
# bytes is ~30 ms on big workbooks — trivial next to the parse it replaces. The
# hash also serves as a content fingerprint, so byte-level changes invalidate
# automatically.
#
# Cache values are stored as JSON-encoded bytes (not Python dicts) so that
# every hit returns a freshly-parsed dict with mutation isolation, without
# paying for `copy.deepcopy` on the way out. On a 50k-cell workbook
# deepcopy was ~450 ms / hit; json round-trip via stdlib was ~96 ms;
# orjson brings that down to ~42 ms — still slower than serving a shared
# reference, but isolation is a hard requirement for cache safety.
#
_XLSX_CACHE_MAX = 8
_XLSX_CACHE: OrderedDict[str, bytes] = OrderedDict()
_XLSX_CACHE_LOCK = Lock()


def _xlsx_cache_disabled() -> bool:
    return os.environ.get("DOXMIND_DISABLE_XLSX_CACHE", "").strip().lower() in {
        "1",
        "true",
        "yes",
        "on",
    }


def _clear_xlsx_cache() -> None:
    """For tests / benchmarks."""
    with _XLSX_CACHE_LOCK:
        _XLSX_CACHE.clear()


@dataclass
class ParseLimits:
    max_sheets: int = MAX_SHEETS
    max_rows: int = MAX_ROWS_PER_SHEET
    max_cols: int = MAX_COLS_PER_SHEET


def parse_workbook(
    xlsx_bytes: bytes,
    *,
    limits: ParseLimits | None = None,
) -> dict[str, Any]:
    """Parse an .xlsx blob into a JSON-serialisable cell model.

    Returns::

        {
          "version": 1,
          "sheets": [
            {
              "id": "sheet-0",
              "name": "Sheet1",
              "index": 0,
              "rowCount": 24,
              "colCount": 8,
              "rowHeights": {"0": 22.0, ...},   # row index -> px (omitted = default)
              "colWidths":  {"0": 96.0, ...},   # col index -> px (omitted = default)
              "merges":   [{"top":1,"left":0,"bottom":2,"right":3}, ...],
              "frozen":   {"row": 1, "col": 0},
              "cells": [
                {
                  "row": 0, "col": 0,
                  "value": "Q1 Revenue",
                  "formula": null,
                  "numberFormat": "General",
                  "style": {"bold": true, "color": "#111111"}
                }, ...
              ]
            }
          ],
          "truncated": {"sheets": false, "rowsBy": {...}, "colsBy": {...}}
        }
    """
    cache_key = _workbook_cache_key(xlsx_bytes)
    cached = _cached_workbook_json(cache_key)
    if cached is not None:
        # JSON round-trip rather than deepcopy: every hit gets a fresh dict
        # (mutation by future callers can't corrupt the entry) while avoiding
        # ~10x the per-cell overhead of CPython's copy.deepcopy on a
        # deeply-nested DTO. orjson roughly halves the decode cost again on
        # multi-MB DTOs.
        return _cache_decode(cached)

    parse_limits = limits or ParseLimits()
    result = _parse_workbook_uncached(xlsx_bytes, parse_limits)
    _store_workbook_json(cache_key, _cache_encode(result))
    return result


def parse_workbook_json_bytes(
    xlsx_bytes: bytes,
    *,
    limits: ParseLimits | None = None,
) -> bytes:
    """Parse an .xlsx blob and return its cached JSON representation."""
    cache_key = _workbook_cache_key(xlsx_bytes)
    cached = _cached_workbook_json(cache_key)
    if cached is not None:
        return cached

    parse_limits = limits or ParseLimits()
    encoded = _cache_encode(_parse_workbook_uncached(xlsx_bytes, parse_limits))
    _store_workbook_json(cache_key, encoded)
    return encoded


def parse_csv_workbook_json_bytes(
    csv_bytes: bytes,
    *,
    limits: ParseLimits | None = None,
) -> bytes:
    """Parse a CSV blob into the same single-sheet JSON model as an .xlsx."""
    if not csv_bytes:
        raise ValueError("empty csv body")
    cache_key = _csv_cache_key(csv_bytes)
    cached = _cached_workbook_json(cache_key)
    if cached is not None:
        return cached

    parse_limits = limits or ParseLimits()
    encoded = _cache_encode(_parse_csv_uncached(csv_bytes, parse_limits))
    _store_workbook_json(cache_key, encoded)
    return encoded


def _workbook_cache_key(xlsx_bytes: bytes) -> str | None:
    if not xlsx_bytes:
        raise ValueError("empty xlsx body")
    if _xlsx_cache_disabled():
        return None
    with perf_timed("excel.parse_workbook.hash", bytes=len(xlsx_bytes)):
        return hashlib.sha256(xlsx_bytes).hexdigest()


def _csv_cache_key(csv_bytes: bytes) -> str | None:
    if _xlsx_cache_disabled():
        return None
    with perf_timed("excel.parse_csv.hash", bytes=len(csv_bytes)):
        return "csv:" + hashlib.sha256(csv_bytes).hexdigest()


def _parse_csv_uncached(csv_bytes: bytes, parse_limits: ParseLimits) -> dict[str, Any]:
    with perf_timed("excel.parse_csv.total", bytes=len(csv_bytes)):
        wb = _workbook_from_csv_bytes(csv_bytes, parse_limits)
        try:
            sheet = wb.active
            assert sheet is not None
            sheet_dto, sheet_truncations = _parse_sheet(
                sheet,
                value_sheet_loader=None,
                index=0,
                limits=parse_limits,
            )
        finally:
            wb.close()

    return {
        "version": 1,
        "sheets": [sheet_dto],
        "truncated": {
            "sheets": False,
            "rowsBy": {"sheet-0": True} if sheet_truncations.get("rows") else {},
            "colsBy": {"sheet-0": True} if sheet_truncations.get("cols") else {},
        },
    }


def _workbook_from_csv_bytes(csv_bytes: bytes, limits: ParseLimits) -> Workbook:
    try:
        text = csv_bytes.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("failed to parse csv: expected UTF-8 text") from exc

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    raw_rows = 0
    raw_cols = 0
    try:
        reader = csv.reader(io.StringIO(text))
        for raw_rows, row in enumerate(reader, start=1):
            raw_cols = max(raw_cols, len(row))
            if raw_rows > limits.max_rows:
                continue
            for col_index, value in enumerate(row[: limits.max_cols], start=1):
                if value != "":
                    ws.cell(row=raw_rows, column=col_index, value=value)
    except csv.Error as exc:
        raise ValueError(f"failed to parse csv: {exc}") from exc
    if raw_rows > limits.max_rows:
        ws.cell(row=limits.max_rows + 1, column=1, value=None)
    if raw_cols > limits.max_cols:
        ws.cell(row=1, column=limits.max_cols + 1, value=None)
    return wb


def _cached_workbook_json(cache_key: str | None) -> bytes | None:
    if cache_key is None:
        return None
    with _XLSX_CACHE_LOCK:
        cached = _XLSX_CACHE.get(cache_key)
        if cached is None:
            return None
        _XLSX_CACHE.move_to_end(cache_key)
        perf_record("excel.parse_workbook.cache_hit")
        return cached


def _store_workbook_json(cache_key: str | None, encoded: bytes) -> None:
    if cache_key is None:
        return
    with _XLSX_CACHE_LOCK:
        _XLSX_CACHE[cache_key] = encoded
        _XLSX_CACHE.move_to_end(cache_key)
        while len(_XLSX_CACHE) > _XLSX_CACHE_MAX:
            _XLSX_CACHE.popitem(last=False)


def _parse_workbook_uncached(xlsx_bytes: bytes, parse_limits: ParseLimits) -> dict[str, Any]:
    with perf_timed("excel.parse_workbook.total", bytes=len(xlsx_bytes)) as total_span:
        wb_formulas: Workbook | None = None
        wb_values: Workbook | None = None
        try:
            # data_only=False so we keep the raw formulas. Frontend gets a
            # lazy second pass with data_only=True only if formula cells need
            # cached values. Plain data tables avoid a full second workbook
            # load on cold open.
            with perf_timed("excel.load_data_only_false"):
                wb_formulas = load_workbook(io.BytesIO(xlsx_bytes), data_only=False, read_only=True)
        except Exception as exc:
            # openpyxl raises a grab-bag of exception types on malformed files.
            raise ValueError(f"failed to parse xlsx: {exc}") from exc

        def value_sheet_for(name: str) -> Worksheet | None:
            nonlocal wb_values
            if wb_values is None:
                with perf_timed("excel.load_data_only_true"):
                    wb_values = load_workbook(
                        io.BytesIO(xlsx_bytes), data_only=True, read_only=True
                    )
            return wb_values[name] if name in wb_values.sheetnames else None

        sheets: list[dict[str, Any]] = []
        truncated_sheets = len(wb_formulas.sheetnames) > parse_limits.max_sheets
        rows_truncated: dict[str, bool] = {}
        cols_truncated: dict[str, bool] = {}
        total_span["sheet_count"] = min(len(wb_formulas.sheetnames), parse_limits.max_sheets)

        try:
            with perf_timed("excel.parse_sheets") as sheets_span:
                cell_total = 0
                for index, name in enumerate(wb_formulas.sheetnames):
                    if index >= parse_limits.max_sheets:
                        break
                    formula_sheet = wb_formulas[name]
                    sheet_dto, sheet_truncations = _parse_sheet(
                        formula_sheet,
                        value_sheet_loader=lambda sheet_name=name: value_sheet_for(sheet_name),
                        index=index,
                        limits=parse_limits,
                    )
                    cell_total += len(sheet_dto.get("cells", []))
                    sheets.append(sheet_dto)
                    if sheet_truncations.get("rows"):
                        rows_truncated[sheet_dto["id"]] = True
                    if sheet_truncations.get("cols"):
                        cols_truncated[sheet_dto["id"]] = True
                sheets_span["cells"] = cell_total
            total_span["cells"] = cell_total
        except Exception as exc:
            raise ValueError(f"failed to parse xlsx: {exc}") from exc
        finally:
            if wb_formulas is not None:
                wb_formulas.close()
            if wb_values is not None:
                wb_values.close()

    result = {
        "version": 1,
        "sheets": sheets,
        "truncated": {
            "sheets": truncated_sheets,
            "rowsBy": rows_truncated,
            "colsBy": cols_truncated,
        },
    }

    return result


def _parse_sheet(
    formula_sheet: Worksheet,
    value_sheet_loader: Callable[[], Worksheet | None] | None = None,
    *,
    index: int,
    limits: ParseLimits,
) -> tuple[dict[str, Any], dict[str, bool]]:
    sheet_id = f"sheet-{index}"
    cells: list[dict[str, Any]] = []
    raw_rows_hint = formula_sheet.max_row or 0
    raw_cols_hint = formula_sheet.max_column or 0
    raw_rows = raw_rows_hint
    raw_cols = raw_cols_hint
    layout: (
        tuple[
            dict[str, float],
            dict[str, float],
            list[dict[str, int]],
            dict[str, int],
            dict[tuple[int, int], str],
        ]
        | None
    ) = None

    if hasattr(formula_sheet, "_cells"):
        row_count = min(raw_rows, limits.max_rows)
        col_count = min(raw_cols, limits.max_cols)
        materialized: list[Cell] = []
        formula_coords: set[tuple[int, int]] = set()
        for cell in formula_sheet._cells.values():
            if cell.row < 1 or cell.row > row_count:
                continue
            if cell.column < 1 or cell.column > col_count:
                continue
            materialized.append(cell)
            raw = cell.value
            if isinstance(raw, str) and raw.startswith("="):
                formula_coords.add((cell.row - 1, cell.column - 1))
        value_sheet = value_sheet_loader() if formula_coords and value_sheet_loader else None
        cached_formula_values = _cached_formula_values(value_sheet, formula_coords)

        for cell in sorted(materialized, key=lambda item: (item.row, item.column)):
            row_idx = cell.row - 1
            col_idx = cell.column - 1
            raw = cell.value
            is_formula = isinstance(raw, str) and raw.startswith("=")
            if (
                raw is not None
                and not is_formula
                and not cell.has_style
                and cell.number_format == "General"
                and cell.hyperlink is None
            ):
                cells.append(
                    {
                        "row": row_idx,
                        "col": col_idx,
                        "value": (
                            raw
                            if isinstance(raw, (str, int, float, bool))
                            else _coerce_jsonable(raw)
                        ),
                        "formula": None,
                    }
                )
                continue
            dto = _cell_to_dto(
                cell,
                cached_value=(
                    cached_formula_values.get((row_idx, col_idx)) if is_formula else None
                ),
                row=row_idx,
                col=col_idx,
            )
            if dto is not None:
                cells.append(dto)
    else:
        max_col_arg = min(raw_cols_hint or limits.max_cols, limits.max_cols)
        materialized: list[Any] = []
        formula_coords: set[tuple[int, int]] = set()
        for row in formula_sheet.iter_rows(
            min_row=1,
            max_row=limits.max_rows,
            min_col=1,
            max_col=max_col_arg,
        ):
            for cell in row:
                row_number = getattr(cell, "row", None)
                column_number = getattr(cell, "column", None)
                if row_number is None or column_number is None:
                    continue
                raw_rows = max(raw_rows, int(row_number))
                raw_cols = max(raw_cols, int(column_number))
                row_idx = int(row_number) - 1
                col_idx = int(column_number) - 1
                materialized.append(cell)
                raw = cell.value
                if isinstance(raw, str) and raw.startswith("="):
                    formula_coords.add((row_idx, col_idx))

        row_count = min(raw_rows, limits.max_rows)
        col_count = min(raw_cols, limits.max_cols)
        layout = _sheet_layout(
            formula_sheet,
            row_count=row_count,
            col_count=col_count,
        )
        read_only_hyperlinks = layout[4]

        cached_formula_values = _cached_formula_values_from_formula_xml(
            formula_sheet,
            formula_coords,
        )
        if len(cached_formula_values) < len(formula_coords):
            value_sheet = value_sheet_loader() if formula_coords and value_sheet_loader else None
            cached_formula_values = {
                **_cached_formula_values(value_sheet, formula_coords),
                **cached_formula_values,
            }

        for cell in materialized:
            row_idx = int(cell.row) - 1
            col_idx = int(cell.column) - 1
            if row_idx >= row_count or col_idx >= col_count:
                continue
            raw = cell.value
            is_formula = isinstance(raw, str) and raw.startswith("=")
            hyperlink = read_only_hyperlinks.get((row_idx, col_idx))
            if (
                raw is not None
                and not is_formula
                and not getattr(cell, "has_style", False)
                and getattr(cell, "number_format", "General") == "General"
                and hyperlink is None
            ):
                cells.append(
                    {
                        "row": row_idx,
                        "col": col_idx,
                        "value": (
                            raw
                            if isinstance(raw, (str, int, float, bool))
                            else _coerce_jsonable(raw)
                        ),
                        "formula": None,
                    }
                )
                continue
            dto = _cell_to_dto(
                cell,
                cached_value=(
                    cached_formula_values.get((row_idx, col_idx)) if is_formula else None
                ),
                hyperlink=hyperlink,
                row=row_idx,
                col=col_idx,
            )
            if dto is not None:
                cells.append(dto)

    row_count = min(raw_rows, limits.max_rows)
    col_count = min(raw_cols, limits.max_cols)

    if layout is None:
        layout = _sheet_layout(formula_sheet, row_count=row_count, col_count=col_count)
    row_heights, col_widths, merges, frozen, _hyperlinks = layout

    return (
        {
            "id": sheet_id,
            "name": formula_sheet.title,
            "index": index,
            "rowCount": row_count,
            "colCount": col_count,
            "rowHeights": row_heights,
            "colWidths": col_widths,
            "merges": merges,
            "frozen": frozen,
            "cells": cells,
        },
        {
            "rows": raw_rows > row_count,
            "cols": raw_cols > col_count,
        },
    )


def _sheet_layout(
    sheet: Any,
    *,
    row_count: int,
    col_count: int,
) -> tuple[
    dict[str, float],
    dict[str, float],
    list[dict[str, int]],
    dict[str, int],
    dict[tuple[int, int], str],
]:
    if hasattr(sheet, "row_dimensions") and hasattr(sheet, "column_dimensions"):
        row_heights: dict[str, float] = {}
        for row_idx, dim in (sheet.row_dimensions or {}).items():
            if dim.height is None:
                continue
            # openpyxl indexes rows from 1; we serialise zero-based.
            if not isinstance(row_idx, int):
                continue
            if row_idx < 1 or row_idx > row_count:
                continue
            row_heights[str(row_idx - 1)] = float(dim.height)

        col_widths: dict[str, float] = {}
        for col_letter, dim in (sheet.column_dimensions or {}).items():
            if dim.width is None:
                continue
            try:
                zero_based = _col_letter_to_index(col_letter)
            except ValueError:
                continue
            if zero_based >= col_count:
                continue
            col_widths[str(zero_based)] = float(dim.width)

        merges: list[dict[str, int]] = []
        for merged_range in sheet.merged_cells.ranges:
            merges.append(
                {
                    "top": merged_range.min_row - 1,
                    "left": merged_range.min_col - 1,
                    "bottom": merged_range.max_row - 1,
                    "right": merged_range.max_col - 1,
                }
            )
        return row_heights, col_widths, merges, _frozen_panes(sheet), {}

    return _read_only_sheet_layout(sheet, row_count=row_count, col_count=col_count)


def _read_only_sheet_layout(
    sheet: Any,
    *,
    row_count: int,
    col_count: int,
) -> tuple[
    dict[str, float],
    dict[str, float],
    list[dict[str, int]],
    dict[str, int],
    dict[tuple[int, int], str],
]:
    archive = getattr(getattr(sheet, "parent", None), "_archive", None)
    worksheet_path = getattr(sheet, "_worksheet_path", None)
    if archive is None or worksheet_path is None:
        return {}, {}, [], {"row": 0, "col": 0}, {}

    row_heights: dict[str, float] = {}
    col_widths: dict[str, float] = {}
    merges: list[dict[str, int]] = []
    frozen = {"row": 0, "col": 0}
    hyperlinks: dict[tuple[int, int], str] = {}

    payload = archive.read(worksheet_path)
    if not any(marker in payload for marker in _SHEET_LAYOUT_MARKERS):
        return row_heights, col_widths, merges, frozen, hyperlinks

    rels = _read_sheet_relationships(archive, worksheet_path)

    with io.BytesIO(payload) as fh:
        for _event, elem in ET.iterparse(fh, events=("end",)):
            tag = elem.tag.rsplit("}", 1)[-1]
            if tag == "row":
                row_ref = elem.get("r")
                height = elem.get("ht")
                if row_ref and height:
                    try:
                        row_idx = int(row_ref)
                        if 1 <= row_idx <= row_count:
                            row_heights[str(row_idx - 1)] = float(height)
                    except ValueError:
                        pass
            elif tag == "col":
                width = elem.get("width")
                if width:
                    try:
                        min_col = int(elem.get("min", "1"))
                        max_col = int(elem.get("max", str(min_col)))
                        parsed_width = float(width)
                    except ValueError:
                        elem.clear()
                        continue
                    for col_idx in range(max(min_col, 1), min(max_col, col_count) + 1):
                        col_widths[str(col_idx - 1)] = parsed_width
            elif tag == "mergeCell":
                ref = elem.get("ref")
                if ref:
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(ref)
                    except ValueError:
                        elem.clear()
                        continue
                    if min_row <= row_count and min_col <= col_count:
                        merges.append(
                            {
                                "top": min_row - 1,
                                "left": min_col - 1,
                                "bottom": min(max_row, row_count) - 1,
                                "right": min(max_col, col_count) - 1,
                            }
                        )
            elif tag == "pane":
                top_left = elem.get("topLeftCell")
                if top_left:
                    frozen = _frozen_from_cell_ref(top_left)
                else:
                    try:
                        frozen = {
                            "row": max(int(float(elem.get("ySplit", "0"))), 0),
                            "col": max(int(float(elem.get("xSplit", "0"))), 0),
                        }
                    except ValueError:
                        frozen = {"row": 0, "col": 0}
            elif tag == "hyperlink":
                ref = elem.get("ref")
                target = elem.get("location")
                rel_id = elem.get(_REL_ID_ATTR)
                if rel_id and rel_id in rels:
                    target = rels[rel_id]
                if ref and target:
                    _add_hyperlink_refs(hyperlinks, ref, target, row_count, col_count)
            elem.clear()

    return row_heights, col_widths, merges, frozen, hyperlinks


def _read_sheet_relationships(archive: Any, worksheet_path: str) -> dict[str, str]:
    path = PurePosixPath(worksheet_path)
    rels_path = path.parent / "_rels" / f"{path.name}.rels"
    try:
        payload = archive.read(str(rels_path))
    except KeyError:
        return {}

    rels: dict[str, str] = {}
    for _event, elem in ET.iterparse(io.BytesIO(payload), events=("end",)):
        if elem.tag.rsplit("}", 1)[-1] != "Relationship":
            elem.clear()
            continue
        rel_id = elem.get("Id")
        target = elem.get("Target")
        if rel_id and target:
            rels[rel_id] = target
        elem.clear()
    return rels


def _add_hyperlink_refs(
    hyperlinks: dict[tuple[int, int], str],
    ref: str,
    target: str,
    row_count: int,
    col_count: int,
) -> None:
    try:
        min_col, min_row, max_col, max_row = range_boundaries(ref)
    except ValueError:
        return
    for row_idx in range(max(min_row, 1), min(max_row, row_count) + 1):
        for col_idx in range(max(min_col, 1), min(max_col, col_count) + 1):
            hyperlinks[(row_idx - 1, col_idx - 1)] = target


def _cached_formula_values_from_formula_xml(
    formula_sheet: Any,
    formula_coords: set[tuple[int, int]],
) -> dict[tuple[int, int], Any]:
    if not formula_coords:
        return {}
    archive = getattr(getattr(formula_sheet, "parent", None), "_archive", None)
    worksheet_path = getattr(formula_sheet, "_worksheet_path", None)
    if archive is None or worksheet_path is None:
        return {}

    shared_strings = getattr(formula_sheet, "_shared_strings", []) or []
    payload = archive.read(worksheet_path)

    cached: dict[tuple[int, int], Any] = {}
    with io.BytesIO(payload) as fh:
        for _event, elem in ET.iterparse(fh, events=("end",)):
            tag = elem.tag.rsplit("}", 1)[-1]
            if tag != "c":
                continue
            ref = elem.get("r")
            if not ref:
                elem.clear()
                continue
            try:
                col, row, _max_col, _max_row = range_boundaries(ref)
            except ValueError:
                elem.clear()
                continue
            coord = (row - 1, col - 1)
            if coord in formula_coords:
                cached[coord] = _cached_formula_cell_value(elem, shared_strings)
                if len(cached) == len(formula_coords):
                    elem.clear()
                    break
            elem.clear()
    return cached


def _cached_formula_cell_value(elem: ET.Element, shared_strings: list[Any]) -> Any:
    cell_type = elem.get("t")
    value_text: str | None = None
    inline_text: str | None = None
    for child in elem:
        tag = child.tag.rsplit("}", 1)[-1]
        if tag == "v":
            value_text = child.text
        elif tag == "is":
            inline_text = "".join(child.itertext())

    if inline_text is not None:
        return inline_text
    if value_text is None:
        return None
    if cell_type == "s":
        try:
            return shared_strings[int(value_text)]
        except (ValueError, IndexError):
            return value_text
    if cell_type == "b":
        return value_text == "1"
    if cell_type in {"str", "e"}:
        return value_text
    try:
        numeric = float(value_text)
    except ValueError:
        return value_text
    return int(numeric) if numeric.is_integer() else numeric


def _cached_formula_values(
    value_sheet: Any | None,
    formula_coords: set[tuple[int, int]],
) -> dict[tuple[int, int], Any]:
    if value_sheet is None or not formula_coords:
        return {}

    min_row = min(row for row, _ in formula_coords) + 1
    max_row = max(row for row, _ in formula_coords) + 1
    min_col = min(col for _, col in formula_coords) + 1
    max_col = max(col for _, col in formula_coords) + 1
    wanted = formula_coords
    cached: dict[tuple[int, int], Any] = {}

    for row_offset, values in enumerate(
        value_sheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
    ):
        row_idx = min_row - 1 + row_offset
        for col_offset, value in enumerate(values):
            col_idx = min_col - 1 + col_offset
            if (row_idx, col_idx) in wanted:
                cached[(row_idx, col_idx)] = value
    return cached


def _cell_to_dto(
    cell: Any,
    *,
    cached_value: Any,
    hyperlink: str | None = None,
    row: int,
    col: int,
) -> dict[str, Any] | None:
    raw = cell.value
    formula: str | None = None
    value: Any = raw

    if isinstance(raw, str) and raw.startswith("="):
        formula = raw
        value = _formula_cached_value_for_cell(cell, cached_value)

    if value is None and formula is None and not _has_visible_style(cell, hyperlink=hyperlink):
        return None

    style = _cell_style(cell, hyperlink=hyperlink)
    dto: dict[str, Any] = {
        "row": row,
        "col": col,
        "value": _coerce_jsonable(value),
        "formula": formula,
    }
    number_format = getattr(cell, "number_format", None)
    if number_format and number_format != "General":
        dto["numberFormat"] = number_format
    if style:
        dto["style"] = style
    return dto


def _formula_cached_value_for_cell(cell: Any, cached_value: Any) -> Any:
    if (
        cached_value is None
        or isinstance(cached_value, bool)
        or not isinstance(cached_value, (int, float))
    ):
        return cached_value
    number_format = getattr(cell, "number_format", None)
    if not number_format or not is_date_format(number_format):
        return cached_value
    workbook = getattr(getattr(cell, "parent", None), "parent", None)
    try:
        epoch = getattr(workbook, "epoch", None)
        return (
            from_excel(cached_value, epoch=epoch) if epoch is not None else from_excel(cached_value)
        )
    except Exception:
        return cached_value


def _has_visible_style(cell: Any, *, hyperlink: str | None = None) -> bool:
    """True when a blank cell still carries paint we want to preserve."""
    if hyperlink:
        return True
    if getattr(cell, "has_style", False):
        font = getattr(cell, "font", None)
        if font is not None and (font.bold or font.italic or font.underline):
            return True
        if font is not None and font.color and font.color.rgb:
            return True
        fill = getattr(cell, "fill", None)
        if fill is not None and fill.fgColor and getattr(fill.fgColor, "rgb", None):
            rgb = fill.fgColor.rgb
            if isinstance(rgb, str) and rgb not in {"00000000", "FFFFFFFF"}:
                return True
    return False


def _cell_style(cell: Any, *, hyperlink: str | None = None) -> dict[str, Any]:
    style: dict[str, Any] = {}
    if not getattr(cell, "has_style", False):
        if hyperlink:
            style["hyperlink"] = hyperlink
        return style
    font = getattr(cell, "font", None)
    if font is not None:
        if font.bold:
            style["bold"] = True
        if font.italic:
            style["italic"] = True
        if font.underline and font.underline != "none":
            style["underline"] = True
        if font.strike:
            style["strikethrough"] = True
        color = _color_to_hex(font.color)
        if color:
            style["color"] = color
        if font.size:
            style["fontSize"] = float(font.size)
        if font.name:
            style["fontFamily"] = font.name
    fill = getattr(cell, "fill", None)
    if fill is not None and fill.fgColor:
        bg = _color_to_hex(fill.fgColor)
        if bg and bg.lower() not in {"#000000", "#ffffff"}:
            style["background"] = bg
    alignment = getattr(cell, "alignment", None)
    if alignment is not None:
        if alignment.horizontal in {"left", "center", "right"}:
            style["textAlign"] = alignment.horizontal
        if alignment.vertical in {"top", "center", "bottom"}:
            style["verticalAlign"] = (
                "middle" if alignment.vertical == "center" else alignment.vertical
            )
        if alignment.wrap_text:
            style["wrapText"] = True
        rotation = alignment.text_rotation
        if isinstance(rotation, (int, float)) and rotation:
            # Map openpyxl's split domain back to a single signed degree
            # value the frontend understands (0–90 up, -1 to -90 down).
            deg = int(rotation)
            if 0 < deg <= 90:
                style["rotation"] = deg
            elif 90 < deg <= 180:
                style["rotation"] = -(deg - 90)
    cell_hyperlink = getattr(cell, "hyperlink", None)
    if hyperlink:
        style["hyperlink"] = hyperlink
    elif cell_hyperlink is not None:
        target = getattr(cell_hyperlink, "target", None) or getattr(cell_hyperlink, "ref", None)
        if isinstance(target, str) and target:
            style["hyperlink"] = target
    border = _border_to_dict(getattr(cell, "border", None))
    if border:
        style["border"] = border
    return style


_OPENPYXL_BORDER_STYLES = {
    "thin": "thin",
    "hair": "thin",
    "medium": "medium",
    "thick": "thick",
    "double": "double",
    "dashed": "dashed",
    "mediumDashed": "dashed",
    "dotted": "dotted",
    "dashDot": "dashed",
    "mediumDashDot": "dashed",
    "dashDotDot": "dashed",
    "mediumDashDotDot": "dashed",
    "slantDashDot": "dashed",
}


def _side_to_dict(side: Any) -> dict[str, Any] | None:
    if side is None:
        return None
    style_name = getattr(side, "style", None)
    if not style_name:
        return None
    mapped = _OPENPYXL_BORDER_STYLES.get(style_name)
    if not mapped:
        return None
    out: dict[str, Any] = {"style": mapped}
    color = _color_to_hex(getattr(side, "color", None))
    if color:
        out["color"] = color
    return out


def _border_to_dict(border: Any) -> dict[str, Any] | None:
    if border is None:
        return None
    out: dict[str, Any] = {}
    for side_name in ("top", "right", "bottom", "left"):
        side = _side_to_dict(getattr(border, side_name, None))
        if side:
            out[side_name] = side
    return out or None


def _color_to_hex(color: Any) -> str | None:
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if not isinstance(rgb, str) or len(rgb) < 6:
        return None
    # openpyxl emits "AARRGGBB" — strip alpha.
    hex_part = rgb[-6:]
    return f"#{hex_part.lower()}"


def _frozen_panes(sheet: Worksheet) -> dict[str, int]:
    cell_ref = sheet.freeze_panes
    if not cell_ref:
        return {"row": 0, "col": 0}
    return _frozen_from_cell_ref(cell_ref)


def _frozen_from_cell_ref(cell_ref: str) -> dict[str, int]:
    try:
        col_letters = "".join(ch for ch in cell_ref if ch.isalpha())
        row_digits = "".join(ch for ch in cell_ref if ch.isdigit())
        col = _col_letter_to_index(col_letters) if col_letters else 0
        row = int(row_digits) - 1 if row_digits else 0
        return {"row": max(row, 0), "col": max(col, 0)}
    except ValueError:
        return {"row": 0, "col": 0}


def _col_letter_to_index(letters: str) -> int:
    letters = letters.strip().upper()
    if not letters or not letters.isalpha():
        raise ValueError(f"invalid column letter: {letters!r}")
    total = 0
    for ch in letters:
        total = total * 26 + (ord(ch) - ord("A") + 1)
    return total - 1


def _coerce_jsonable(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    # Datetimes, Decimals, etc. — fall back to string. The frontend doesn't
    # need full type fidelity for the spike; rich types come back from the
    # next phase when we wire openpyxl numberFormat → renderer.
    try:
        return str(value)
    except Exception:
        return None
