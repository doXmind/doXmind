"""S4 (ADR 0010): binary read — convert (CLI) + read_pdf/read_excel (MCP)."""

import json
from pathlib import Path

import fitz  # PyMuPDF
import pytest
from openpyxl import Workbook
from typer.testing import CliRunner

from cli.__main__ import app
from core.convert import convert_excel, convert_pdf
from doxmind_mcp import server

runner = CliRunner()


def _make_pdf(path: Path) -> Path:
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((72, 72), "Hello PDF block")
    path.write_bytes(doc.tobytes())
    doc.close()
    return path


def _make_xlsx(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Hello"
    ws["B1"] = 42
    wb.save(path)
    return path


def test_convert_pdf_returns_pages(tmp_path):
    result = convert_pdf(_make_pdf(tmp_path / "doc.pdf"))
    assert result["pageCount"] == 1
    assert result["pages"]


def test_convert_excel_returns_sheets(tmp_path):
    result = convert_excel(_make_xlsx(tmp_path / "book.xlsx"))
    assert result["sheets"]


def test_cli_convert_pdf(tmp_path):
    pdf = _make_pdf(tmp_path / "doc.pdf")
    res = runner.invoke(app, ["convert", str(pdf)])
    assert res.exit_code == 0, res.output
    assert json.loads(res.stdout)["pageCount"] == 1


def test_cli_convert_rejects_unsupported(tmp_path):
    md = tmp_path / "note.md"
    md.write_text("hi", encoding="utf-8")
    res = runner.invoke(app, ["convert", str(md)])
    assert res.exit_code != 0


def test_mcp_read_pdf_and_excel_confined(tmp_path, monkeypatch):
    _make_pdf(tmp_path / "doc.pdf")
    _make_xlsx(tmp_path / "book.xlsx")
    monkeypatch.setenv("DOXMIND_WORKSPACE_ROOT", str(tmp_path))
    assert server.read_pdf("doc.pdf")["pageCount"] == 1
    assert server.read_excel("book.xlsx")["sheets"]
    with pytest.raises(ValueError):
        server.read_pdf("../escape.pdf")
