"""Regression coverage for read-only Excel attachment conversion."""

from __future__ import annotations

import io
import zipfile
from xml.etree import ElementTree as ET

from openpyxl import Workbook

from services.excel_workbook import parse_workbook


def divergent_iterator_workbook() -> bytes:
    """Return sparse XLSX bytes whose read-only value stream used to diverge."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Header"
    ws["A2"] = "Dangling"

    source = io.BytesIO()
    wb.save(source)

    malformed = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(source.getvalue())) as zin:
        with zipfile.ZipFile(malformed, "w") as zout:
            for item in zin.infolist():
                payload = zin.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    payload = payload.decode().replace('r="A2"', 'r="A3"').encode()
                zout.writestr(item, payload)
    return malformed.getvalue()


def cached_formula_workbook() -> bytes:
    """Return XLSX bytes with a formula and an embedded cached result."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["B1"] = 2
    ws["C1"] = "=A1+B1"

    source = io.BytesIO()
    wb.save(source)

    patched = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(source.getvalue())) as zin:
        with zipfile.ZipFile(patched, "w") as zout:
            for item in zin.infolist():
                payload = zin.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    root = ET.fromstring(payload)
                    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
                    cell = root.find(f".//{namespace}c[@r='C1']")
                    assert cell is not None
                    for value_node in cell.findall(f"{namespace}v"):
                        cell.remove(value_node)
                    value_node = ET.SubElement(cell, f"{namespace}v")
                    value_node.text = "3"
                    payload = ET.tostring(root, encoding="utf-8")
                zout.writestr(item, payload)
    return patched.getvalue()


def test_parse_workbook_accepts_sparse_xlsx_without_value_stream_divergence():
    sheet = parse_workbook(divergent_iterator_workbook())["sheets"][0]

    assert sheet["rowCount"] == 3
    assert sheet["cells"] == [
        {"row": 0, "col": 0, "value": "Header", "formula": None},
        {"row": 2, "col": 0, "value": "Dangling", "formula": None},
    ]


def test_parse_workbook_reads_cached_values_only_for_formula_cells():
    cells = parse_workbook(cached_formula_workbook())["sheets"][0]["cells"]

    assert cells[2] == {"row": 0, "col": 2, "value": 3, "formula": "=A1+B1"}
