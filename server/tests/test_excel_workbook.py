"""Tests for the openpyxl-backed workbook parse path + its LRU cache.

The full parse_workbook contract is covered indirectly by the preflight
workflow tests (`test_excel_preflight_workflows.py`); this file focuses on
the cache invariants that landed alongside the perf pass.
"""

from __future__ import annotations

import io
import json
import zipfile
from xml.etree import ElementTree as ET

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from services.excel_workbook import (
    parse_csv_workbook_json_bytes,
    parse_workbook,
    parse_workbook_json_bytes,
)


def _build_xlsx(rows: int = 3, cols: int = 3) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=f"R{r}C{c}")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_empty_xlsx() -> bytes:
    wb = Workbook()
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_formula_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = 1
    ws["B1"] = "=A1"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_recovery_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Source"
    ws["A1"] = "ORIGINAL"
    ws.merge_cells("C1:D1")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_cached_date_formula_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = 1
    ws["B1"] = "=DATE(2026,5,28)"
    ws["B1"].number_format = "yyyy-mm-dd"
    buffer = io.BytesIO()
    wb.save(buffer)

    patched = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(buffer.getvalue())) as zin:
        with zipfile.ZipFile(patched, "w") as zout:
            for item in zin.infolist():
                payload = zin.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    root = ET.fromstring(payload)
                    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
                    cell = root.find(f".//{namespace}c[@r='B1']")
                    assert cell is not None
                    for value_node in cell.findall(f"{namespace}v"):
                        cell.remove(value_node)
                    value_node = ET.SubElement(cell, f"{namespace}v")
                    value_node.text = "46170"
                    payload = ET.tostring(root, encoding="utf-8")
                zout.writestr(item, payload)
    return patched.getvalue()


def _build_formula_xlsx_with_out_of_range_empty_cache() -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = 1
    ws["B1"] = "=A1"
    ws["A6000"] = "=A1"
    buffer = io.BytesIO()
    wb.save(buffer)

    patched = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(buffer.getvalue())) as zin:
        with zipfile.ZipFile(patched, "w") as zout:
            for item in zin.infolist():
                payload = zin.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    root = ET.fromstring(payload)
                    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
                    cell = root.find(f".//{namespace}c[@r='B1']")
                    assert cell is not None
                    for value_node in cell.findall(f"{namespace}v"):
                        cell.remove(value_node)
                    value_node = ET.SubElement(cell, f"{namespace}v")
                    value_node.text = "1"
                    payload = ET.tostring(root, encoding="utf-8")
                zout.writestr(item, payload)
    return patched.getvalue()


def _build_formatted_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Formatted"
    ws["A1"] = "Header"
    ws["A1"].font = ws["A1"].font.copy(bold=True)
    ws["A2"] = "Link"
    ws["A2"].hyperlink = "https://example.com"
    ws["B2"] = 1234
    ws["B2"].number_format = '"$"#,##0.00'
    ws["B2"].fill = ws["B2"].fill.copy(fill_type="solid", fgColor="FFF2CC")
    ws.row_dimensions[1].height = 24
    ws.column_dimensions["B"].width = 18
    ws.merge_cells("C3:D4")
    ws.freeze_panes = "B2"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_parse_workbook_read_only_path_preserves_layout_and_styles() -> None:
    from services.excel_workbook import _clear_xlsx_cache

    _clear_xlsx_cache()
    sheet = parse_workbook(_build_formatted_xlsx())["sheets"][0]

    assert sheet["rowHeights"] == {"0": 24.0}
    assert sheet["colWidths"] == {"1": 18.0}
    assert sheet["merges"] == [{"top": 2, "left": 2, "bottom": 3, "right": 3}]
    assert sheet["frozen"] == {"row": 1, "col": 1}
    assert sheet["cells"][0]["style"]["bold"] is True
    link_cell = next(cell for cell in sheet["cells"] if cell["row"] == 1 and cell["col"] == 0)
    assert link_cell["style"]["hyperlink"] == "https://example.com"
    money_cell = next(cell for cell in sheet["cells"] if cell["row"] == 1 and cell["col"] == 1)
    assert money_cell["numberFormat"] == '"$"#,##0.00'
    assert money_cell["style"]["background"] == "#fff2cc"


def test_parse_workbook_skips_data_only_load_without_formulas(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import services.excel_workbook as workbook_module
    from services.excel_workbook import _clear_xlsx_cache

    calls: list[tuple[bool | None, bool | None]] = []
    real_load_workbook = workbook_module.load_workbook

    def spy_load_workbook(*args, **kwargs):
        calls.append((kwargs.get("data_only"), kwargs.get("read_only")))
        return real_load_workbook(*args, **kwargs)

    monkeypatch.setattr(workbook_module, "load_workbook", spy_load_workbook)
    _clear_xlsx_cache()

    parse_workbook(_build_xlsx(rows=2, cols=2))

    assert calls == [(False, True)]


def test_parse_workbook_reads_formula_cached_values_from_formula_xml(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import services.excel_workbook as workbook_module
    from services.excel_workbook import _clear_xlsx_cache

    calls: list[tuple[bool | None, bool | None]] = []
    real_load_workbook = workbook_module.load_workbook

    def spy_load_workbook(*args, **kwargs):
        calls.append((kwargs.get("data_only"), kwargs.get("read_only")))
        return real_load_workbook(*args, **kwargs)

    monkeypatch.setattr(workbook_module, "load_workbook", spy_load_workbook)
    _clear_xlsx_cache()

    parse_workbook(_build_formula_xlsx())

    assert calls == [(False, True)]


def test_parse_workbook_preserves_cached_formula_date_values() -> None:
    from services.excel_workbook import _clear_xlsx_cache

    _clear_xlsx_cache()
    sheet = parse_workbook(_build_cached_date_formula_xlsx())["sheets"][0]
    date_cell = next(cell for cell in sheet["cells"] if cell["row"] == 0 and cell["col"] == 1)

    assert date_cell["formula"] == "=DATE(2026,5,28)"
    assert date_cell["value"] == "2026-05-28 00:00:00"


def test_parse_workbook_ignores_out_of_range_empty_formula_cache() -> None:
    from services.excel_workbook import _clear_xlsx_cache

    _clear_xlsx_cache()
    sheet = parse_workbook(_build_formula_xlsx_with_out_of_range_empty_cache())["sheets"][0]
    formula_cell = next(cell for cell in sheet["cells"] if cell["row"] == 0 and cell["col"] == 1)

    assert formula_cell["value"] == 1


def test_parse_cache_hit_returns_clone_so_caller_mutation_doesnt_leak() -> None:
    from services.excel_workbook import _clear_xlsx_cache

    _clear_xlsx_cache()
    xlsx = _build_xlsx()

    first = parse_workbook(xlsx)
    first["sheets"][0]["cells"] = []  # simulate a downstream filter / mutation

    second = parse_workbook(xlsx)
    assert len(second["sheets"][0]["cells"]) > 0, "cache was corrupted by caller mutation"


def test_parse_cache_invalidated_by_byte_flip() -> None:
    from services.excel_workbook import _clear_xlsx_cache

    _clear_xlsx_cache()
    xlsx_a = _build_xlsx(rows=2, cols=2)
    xlsx_b = _build_xlsx(rows=4, cols=2)

    result_a = parse_workbook(xlsx_a)
    result_b = parse_workbook(xlsx_b)

    # Different byte content → different cache key → distinct results.
    assert result_a["sheets"][0]["rowCount"] == 2
    assert result_b["sheets"][0]["rowCount"] == 4


def test_parse_workbook_json_bytes_reuses_wire_cache() -> None:
    from services.excel_workbook import _clear_xlsx_cache

    _clear_xlsx_cache()
    xlsx = _build_xlsx(rows=2, cols=2)

    first = parse_workbook_json_bytes(xlsx)
    second = parse_workbook_json_bytes(xlsx)

    assert second == first
    assert json.loads(second)["sheets"][0]["rowCount"] == 2


def test_parse_csv_workbook_json_bytes_returns_single_sheet_grid() -> None:
    from services.excel_workbook import _clear_xlsx_cache

    _clear_xlsx_cache()
    result = json.loads(
        parse_csv_workbook_json_bytes(
            b'name,amount,note\nAda,12,"keeps, comma"\nGrace,,blank amount\n'
        )
    )

    sheet = result["sheets"][0]
    assert sheet["name"] == "Sheet1"
    assert sheet["rowCount"] == 3
    assert sheet["colCount"] == 3
    assert sheet["cells"] == [
        {"row": 0, "col": 0, "value": "name", "formula": None},
        {"row": 0, "col": 1, "value": "amount", "formula": None},
        {"row": 0, "col": 2, "value": "note", "formula": None},
        {"row": 1, "col": 0, "value": "Ada", "formula": None},
        {"row": 1, "col": 1, "value": "12", "formula": None},
        {"row": 1, "col": 2, "value": "keeps, comma", "formula": None},
        {"row": 2, "col": 0, "value": "Grace", "formula": None},
        {"row": 2, "col": 2, "value": "blank amount", "formula": None},
    ]
    assert result["truncated"] == {"sheets": False, "rowsBy": {}, "colsBy": {}}


def test_parse_cache_disabled_via_env(monkeypatch: pytest.MonkeyPatch) -> None:
    from services.excel_workbook import _clear_xlsx_cache

    monkeypatch.setenv("DOXMIND_DISABLE_XLSX_CACHE", "1")
    _clear_xlsx_cache()
    xlsx = _build_xlsx()

    first = parse_workbook(xlsx)
    first["sheets"][0]["cells"] = []
    second = parse_workbook(xlsx)
    # With cache off, mutation can't leak — second is a fresh parse.
    assert len(second["sheets"][0]["cells"]) > 0


def test_recovery_export_rejects_cell_targeting_missing_sheet(sync_client) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=1, cols=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps(
                {
                    "version": 1,
                    "cells": {"sheet-9!0,0": {"value": "RECOVERED"}},
                }
            ),
        },
    )

    assert response.status_code == 400
    assert "sheet-9" in response.json()["error"]["message"]


def test_legacy_editor_export_remains_permissive_for_missing_sheet_edits(sync_client) -> None:
    source = _build_xlsx(rows=1, cols=1)
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                source,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "edits": json.dumps(
                {
                    "version": 1,
                    "cells": {"sheet-9!0,0": {"value": "RECOVERED"}},
                }
            )
        },
    )

    assert response.status_code == 200
    assert load_workbook(io.BytesIO(response.content)).active["A1"].value == "R1C1"


@pytest.mark.parametrize(
    "state_patch",
    [
        {"rowHeights": {"sheet-9!0": 24}},
        {"colWidths": {"sheet-9!0": 120}},
        {"frozen": {"sheet-9": {"row": 1, "col": 0}}},
        {"validations": {"sheet-9!0,0": {"type": "list", "values": ["A"]}}},
        {"comments": {"sheet-9!0,0": {"text": "Lost note"}}},
        {
            "conditionalFormats": {
                "sheet-9": [
                    {
                        "id": "lost-rule",
                        "range": {"top": 0, "left": 0, "bottom": 0, "right": 0},
                        "condition": {"kind": "blank"},
                    }
                ]
            }
        },
    ],
)
def test_recovery_export_rejects_every_edit_targeting_a_missing_sheet(
    sync_client, state_patch
) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=1, cols=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps({"version": 1, **state_patch}),
        },
    )

    assert response.status_code == 400
    assert "sheet-9" in response.json()["error"]["message"]


@pytest.mark.parametrize(
    "state",
    [
        {"version": 2, "cells": {"sheet-0!0,0": {"value": "RECOVERED"}}},
        {"version": 1, "mystery": {"userContent": "RECOVERED"}},
        {"version": 1, "activeSheetId": 0, "cells": {}},
        {"version": 1, "cells": []},
        {"version": 1, "rowHeights": []},
        {"version": 1, "colWidths": []},
        {"version": 1, "ops": {}},
        {"version": 1, "workbookOps": {}},
        {"version": 1, "frozen": []},
        {"version": 1, "validations": []},
        {"version": 1, "comments": []},
        {"version": 1, "conditionalFormats": []},
    ],
)
def test_recovery_export_rejects_malformed_top_level_state(sync_client, state) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=1, cols=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"strict_recovery": "true", "edits": json.dumps(state)},
    )

    assert response.status_code == 400


@pytest.mark.parametrize("number", ["NaN", "Infinity", "-Infinity", "1e400", "-1e400"])
def test_excel_export_endpoint_rejects_non_finite_json_numbers(sync_client, number: str) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=1, cols=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": (
                '{"version":1,"cells":{"sheet-0!0,0":{"value":'
                f"{number}"
                "}}}"
            ),
        },
    )

    assert response.status_code == 400


@pytest.mark.parametrize(
    "state_patch",
    [
        {"cells": {"sheet-0!bad": {"value": "RECOVERED"}}},
        {"cells": {"sheet-0!-1,0": {"value": "RECOVERED"}}},
        {"cells": {"sheet-0!0,0": []}},
        {"cells": {"sheet-0!0,0": {}}},
        {"cells": {"sheet-0!0,0": {"value": "RECOVERED", "lost": True}}},
        {"cells": {"sheet-0!0,0": {"value": {"nested": "RECOVERED"}}}},
        {"cells": {"sheet-0!0,0": {"formula": 42}}},
        {"cells": {"sheet-0!0,0": {"numberFormat": 42}}},
        {"cells": {"sheet-0!0,0": {"style": "bold"}}},
        {"cells": {"sheet-0!0,0": {"style": {"bold": "yes"}}}},
        {"cells": {"sheet-0!0,0": {"style": {"textAlign": "diagonal"}}}},
        {"cells": {"sheet-0!0,0": {"style": {"rotation": 1.5}}}},
        {"cells": {"sheet-0!0,0": {"style": {"textOverflow": "clip"}}}},
        {"cells": {"sheet-0!0,0": {"style": {"textOverflow": "overflow"}}}},
        {"cells": {"sheet-0!0,0": {"style": {"background": "not-a-color"}}}},
        {"cells": {"sheet-0!0,0": {"style": {"background": "#gggggg"}}}},
        {"cells": {"sheet-0!0,0": {"style": {"fontSize": 410}}}},
        {"cells": {"sheet-0!0,0": {"style": {"border": {"top": {"style": "hairline"}}}}}},
        {"rowHeights": {"sheet-0!bad": 24}},
        {"rowHeights": {"sheet-0!0": "24"}},
        {"rowHeights": {"sheet-0!0": 410}},
        {"colWidths": {"sheet-0!-1": 120}},
        {"colWidths": {"sheet-0!0": True}},
        {"colWidths": {"sheet-0!0": 256}},
        {"cells": {"sheet-0!0,0": {"value": "x" * 32_768}}},
        {"cells": {"sheet-0!0,0": {"value": "control\x01text"}}},
        {"cells": {"sheet-0!0,0": {"formula": "=" + "x" * 8_192}}},
        {"cells": {"sheet-0!0,0": {"formula": "=control\x01text"}}},
    ],
)
def test_recovery_export_rejects_malformed_cells_and_size_overrides(
    sync_client, state_patch
) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=1, cols=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps({"version": 1, **state_patch}),
        },
    )

    assert response.status_code == 400


def test_recovery_export_rejects_formula_patch_with_non_null_value(sync_client) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=1, cols=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps(
                {
                    "version": 1,
                    "cells": {"sheet-0!0,0": {"formula": "=1+1", "value": "ignored"}},
                }
            ),
        },
    )

    assert response.status_code == 400
    assert "non-null value" in response.json()["error"]["message"]


def test_recovery_export_rejects_conflicting_wrap_style_fields(sync_client) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=1, cols=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps(
                {
                    "version": 1,
                    "cells": {
                        "sheet-0!0,0": {
                            "style": {"wrapText": False, "textOverflow": "wrap"}
                        }
                    },
                }
            ),
        },
    )

    assert response.status_code == 400
    assert "wrapText" in response.json()["error"]["message"]


def test_recovery_export_allows_equivalent_wrap_style_fields(sync_client) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=1, cols=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps(
                {
                    "version": 1,
                    "cells": {
                        "sheet-0!0,0": {
                            "style": {"wrapText": True, "textOverflow": "wrap"}
                        }
                    },
                }
            ),
        },
    )

    assert response.status_code == 200
    cell = load_workbook(io.BytesIO(response.content)).active["A1"]
    assert cell.alignment.wrap_text is True


@pytest.mark.parametrize(
    "cell_patch",
    [
        {"style": {"hyperlink": "https://example.com"}},
        {"value": None, "style": {"hyperlink": "https://example.com"}},
        {"value": "", "style": {"hyperlink": "https://example.com"}},
    ],
)
def test_recovery_export_rejects_hyperlink_when_final_cell_value_is_empty(
    sync_client, cell_patch
) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_empty_xlsx(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps(
                {"version": 1, "cells": {"sheet-0!0,0": cell_patch}}
            ),
        },
    )

    assert response.status_code == 400
    assert "hyperlink" in response.json()["error"]["message"]
    assert "empty cell" in response.json()["error"]["message"]


@pytest.mark.parametrize(
    ("source_bytes", "cell_patch", "expected_value"),
    [
        (
            _build_xlsx(rows=1, cols=1),
            {"style": {"hyperlink": "https://example.com"}},
            "R1C1",
        ),
        (
            _build_empty_xlsx(),
            {"value": "Example", "style": {"hyperlink": "https://example.com"}},
            "Example",
        ),
    ],
)
def test_recovery_export_preserves_hyperlink_when_final_cell_value_is_non_empty(
    sync_client, source_bytes, cell_patch, expected_value
) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                source_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps(
                {
                    "version": 1,
                    "cells": {"sheet-0!0,0": cell_patch},
                }
            ),
        },
    )

    assert response.status_code == 200
    cell = load_workbook(io.BytesIO(response.content)).active["A1"]
    assert cell.value == expected_value
    assert cell.hyperlink.target == "https://example.com"


@pytest.mark.parametrize(
    "state_patch",
    [
        {"workbookOps": ["rename"]},
        {"workbookOps": [{"type": "unknown", "sheetId": "sheet-0"}]},
        {"workbookOps": [{"type": "addSheet", "sheetId": "", "name": "Added"}]},
        {"workbookOps": [{"type": "addSheet", "sheetId": "sheet-0", "name": "Added"}]},
        {"workbookOps": [{"type": "renameSheet", "sheetId": "sheet-0", "name": 42}]},
        {"workbookOps": [{"type": "deleteSheet", "sheetId": "sheet-0"}]},
        {"ops": ["insertRow"]},
        {"ops": [{"type": "unknown", "sheetId": "sheet-0"}]},
        {"ops": [{"type": "insertRow", "sheetId": "sheet-0", "before": "0", "count": 1}]},
        {"ops": [{"type": "deleteCol", "sheetId": "sheet-0", "index": 0, "count": 0}]},
        {
            "ops": [
                {
                    "type": "mergeCells",
                    "sheetId": "sheet-0",
                    "top": 1,
                    "left": 1,
                    "bottom": 0,
                    "right": 1,
                }
            ]
        },
        {
            "ops": [
                {
                    "type": "mergeCells",
                    "sheetId": "sheet-0",
                    "top": 0,
                    "left": 0,
                    "bottom": 0,
                    "right": 0,
                }
            ]
        },
        {
            "ops": [
                {
                    "type": "unmergeCells",
                    "sheetId": "sheet-0",
                    "top": 0,
                    "left": 0,
                    "bottom": 0,
                    "right": 1,
                }
            ]
        },
    ],
)
def test_recovery_export_rejects_malformed_or_unapplied_operations(
    sync_client, state_patch
) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=2, cols=2),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps({"version": 1, **state_patch}),
        },
    )

    assert response.status_code == 400


@pytest.mark.parametrize(
    "operation",
    [
        {"type": "insertRow", "sheetId": "sheet-0", "before": 0, "count": 1},
        {"type": "deleteRow", "sheetId": "sheet-0", "index": 0, "count": 1},
        {"type": "insertCol", "sheetId": "sheet-0", "before": 0, "count": 1},
        {"type": "deleteCol", "sheetId": "sheet-0", "index": 0, "count": 1},
        {
            "type": "mergeCells",
            "sheetId": "sheet-0",
            "top": 0,
            "left": 0,
            "bottom": 0,
            "right": 1,
        },
        {
            "type": "unmergeCells",
            "sheetId": "sheet-0",
            "top": 0,
            "left": 0,
            "bottom": 0,
            "right": 1,
        },
    ],
)
def test_strict_recovery_rejects_structural_ops_that_openpyxl_cannot_replay_exactly(
    sync_client, operation
) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=2, cols=2),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps({"version": 1, "ops": [operation]}),
        },
    )

    assert response.status_code == 400
    assert "structural workbook operations" in response.json()["error"]["message"]


@pytest.mark.parametrize(
    "state_patch",
    [
        {"cells": {"sheet-0!0,3": {"value": "RECOVERED"}}},
        {"comments": {"sheet-0!0,3": {"text": "Recovered note"}}},
        {"validations": {"sheet-0!0,3": {"type": "list", "values": ["A"]}}},
    ],
)
def test_strict_recovery_rejects_non_anchor_merged_cell_targets(
    sync_client, state_patch
) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_recovery_xlsx(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps({"version": 1, **state_patch}),
        },
    )

    assert response.status_code == 400
    assert "non-anchor merged cell" in response.json()["error"]["message"]


@pytest.mark.parametrize(
    "workbook_op",
    [
        {"type": "renameSheet", "sheetId": "sheet-0", "name": "Renamed"},
        {
            "type": "duplicateSheet",
            "sourceSheetId": "sheet-0",
            "sheetId": "sheet-copy",
            "name": "Copy",
        },
        {"type": "deleteSheet", "sheetId": "sheet-0"},
    ],
)
def test_strict_recovery_rejects_lossy_workbook_ops(sync_client, workbook_op) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=2, cols=2),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps({"version": 1, "workbookOps": [workbook_op]}),
        },
    )

    assert response.status_code == 400


@pytest.mark.parametrize(
    "state_patch",
    [
        {"frozen": {"sheet-0": []}},
        {"frozen": {"sheet-0": {"row": 1}}},
        {"frozen": {"sheet-0": {"row": "1", "col": 0}}},
        {"frozen": {"sheet-0": {"row": -1, "col": 0}}},
        {"frozen": {"sheet-0": {"row": 1, "col": 0, "lost": True}}},
        {"validations": {"sheet-0!0,0": []}},
        {"validations": {"sheet-0!0,0": {"type": "number", "values": ["A"]}}},
        {"validations": {"sheet-0!0,0": {"type": "list", "values": []}}},
        {"validations": {"sheet-0!0,0": {"type": "list", "values": ["A", 2]}}},
        {"validations": {"sheet-0!0,0": {"type": "list", "values": ["A"], "lost": True}}},
        {"validations": {"sheet-0!0,0": {"type": "list", "values": ["A" * 254]}}},
        {"validations": {"sheet-0!0,0": {"type": "list", "values": ["A,B"]}}},
        {"comments": {"sheet-0!0,0": []}},
        {"comments": {"sheet-0!0,0": {"text": ""}}},
        {"comments": {"sheet-0!0,0": {"text": "Note", "author": 42}}},
        {"comments": {"sheet-0!0,0": {"text": "Note", "updatedAt": 42}}},
        {
            "comments": {
                "sheet-0!0,0": {"text": "Note", "updatedAt": "not-an-iso-timestamp"}
            }
        },
        {
            "comments": {
                "sheet-0!0,0": {"text": "Note", "updatedAt": "2026-07-20T00:00:00"}
            }
        },
        {"comments": {"sheet-0!0,0": {"text": "x" * 32_768}}},
        {"comments": {"sheet-0!0,0": {"text": "Note", "lost": True}}},
    ],
)
def test_recovery_export_rejects_malformed_freezes_validations_and_comments(
    sync_client, state_patch
) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=2, cols=2),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps({"version": 1, **state_patch}),
        },
    )

    assert response.status_code == 400


@pytest.mark.parametrize(
    "rules",
    [
        {},
        ["rule"],
        [{"id": "missing-fields"}],
        [
            {
                "id": "bad-range",
                "range": {"top": 1, "left": 0, "bottom": 0, "right": 0},
                "condition": {"kind": "blank"},
            }
        ],
        [
            {
                "id": "bad-kind",
                "range": {"top": 0, "left": 0, "bottom": 0, "right": 0},
                "condition": {"kind": "future"},
            }
        ],
        [
            {
                "id": "bad-op",
                "range": {"top": 0, "left": 0, "bottom": 0, "right": 0},
                "condition": {"kind": "cellValue", "op": "contains", "value": 1},
            }
        ],
        [
            {
                "id": "bad-between",
                "range": {"top": 0, "left": 0, "bottom": 0, "right": 0},
                "condition": {"kind": "between", "min": "1", "max": 2},
            }
        ],
        [
            {
                "id": "bad-text",
                "range": {"top": 0, "left": 0, "bottom": 0, "right": 0},
                "condition": {"kind": "containsText", "text": "", "mode": "contains"},
            }
        ],
        [
            {
                "id": "bad-scale",
                "range": {"top": 0, "left": 0, "bottom": 1, "right": 0},
                "condition": {
                    "kind": "colorScale",
                    "min": {"color": "not-a-color"},
                    "max": {"color": "#00ff00"},
                },
            }
        ],
        [
            {
                "id": "bad-style",
                "range": {"top": 0, "left": 0, "bottom": 0, "right": 0},
                "condition": {"kind": "blank"},
                "style": {"fontSize": 12},
            }
        ],
        [
            {
                "id": "false-style",
                "range": {"top": 0, "left": 0, "bottom": 0, "right": 0},
                "condition": {"kind": "blank"},
                "style": {"bold": False},
            }
        ],
        [
            {
                "id": "styled-scale",
                "range": {"top": 0, "left": 0, "bottom": 1, "right": 0},
                "condition": {
                    "kind": "colorScale",
                    "min": {"color": "#ff0000"},
                    "max": {"color": "#00ff00"},
                },
                "style": {"bold": True},
            }
        ],
    ],
)
def test_recovery_export_rejects_malformed_conditional_formats(sync_client, rules) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=2, cols=2),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps({"version": 1, "conditionalFormats": {"sheet-0": rules}}),
        },
    )

    assert response.status_code == 400


def test_recovery_export_applies_every_supported_state_field(sync_client) -> None:
    state = {
        "version": 1,
        "activeSheetId": "sheet-user-recovered",
        "workbookOps": [
            {
                "type": "addSheet",
                "sheetId": "sheet-user-recovered",
                "name": "Recovered",
                "afterSheetId": "sheet-0",
            }
        ],
        "cells": {
            "sheet-0!1,0": {
                "value": "RECOVERED",
                "formula": None,
                "numberFormat": "@",
                "style": {
                    "bold": True,
                    "background": "#ffeeaa",
                    "border": {"bottom": {"style": "thin", "color": "#112233"}},
                },
            },
            "sheet-0!1,1": {"formula": "=1+1", "value": None},
        },
        "rowHeights": {"sheet-0!1": 24},
        "colWidths": {"sheet-0!0": 18},
        "frozen": {"sheet-0": {"row": 1, "col": 1}},
        "validations": {"sheet-0!1,2": {"type": "list", "values": ["A", "B"]}},
        "comments": {
            "sheet-0!1,0": {
                "text": "Recovered note",
                "author": "Ada",
                "updatedAt": "2026-07-20T00:00:00.000Z",
            }
        },
        "conditionalFormats": {
            "sheet-0": [
                {
                    "id": "recovered-rule",
                    "range": {"top": 1, "left": 0, "bottom": 1, "right": 1},
                    "condition": {"kind": "cellValue", "op": "gt", "value": 1},
                    "style": {"bold": True, "background": "#fff2cc"},
                }
            ]
        },
    }

    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_recovery_xlsx(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"strict_recovery": "true", "edits": json.dumps(state)},
    )

    assert response.status_code == 200
    wb = load_workbook(io.BytesIO(response.content), data_only=False)
    assert wb.sheetnames == ["Source", "Recovered"]
    assert wb.active.title == "Recovered"
    ws = wb["Source"]
    assert ws["A2"].value == "RECOVERED"
    assert ws["A2"].number_format == "@"
    assert ws["A2"].font.bold is True
    assert ws["B2"].value == "=1+1"
    assert ws.row_dimensions[2].height == 24
    assert ws.column_dimensions["A"].width == 18
    assert ws.freeze_panes == "B2"
    assert ws["A2"].comment.text == "Recovered note"
    assert ws["A2"].comment.author == "Ada"
    validations = list(ws.data_validations.dataValidation)
    assert len(validations) == 1
    assert str(validations[0].sqref) == "C2"
    assert len(list(ws.conditional_formatting)) == 1
    assert {str(cell_range) for cell_range in ws.merged_cells.ranges} == {"C1:D1"}


def test_recovery_export_preserves_conditional_format_options(sync_client) -> None:
    rules = [
        {
            "id": "exclusive-between",
            "range": {"top": 0, "left": 0, "bottom": 1, "right": 0},
            "condition": {"kind": "between", "min": 1, "max": 2, "inclusive": False},
            "style": {"bold": True},
        },
        {
            "id": "case-sensitive",
            "range": {"top": 0, "left": 1, "bottom": 1, "right": 1},
            "condition": {
                "kind": "containsText",
                "text": "X",
                "mode": "contains",
                "caseSensitive": True,
            },
            "style": {"bold": True},
        },
        {
            "id": "numeric-scale",
            "range": {"top": 0, "left": 2, "bottom": 1, "right": 2},
            "condition": {
                "kind": "colorScale",
                "min": {"value": 0, "color": "#ff0000"},
                "mid": {"value": 50, "color": "#ffff00"},
                "max": {"value": 100, "color": "#00ff00"},
            },
        },
        {
            "id": "string-operand",
            "range": {"top": 0, "left": 0, "bottom": 1, "right": 0},
            "condition": {"kind": "cellValue", "op": "eq", "value": "001"},
        },
        {
            "id": "production-scale",
            "range": {"top": 0, "left": 0, "bottom": 1, "right": 0},
            "condition": {
                "kind": "colorScale",
                "min": {"color": "#ff0000"},
                "mid": {"color": "#ffff00"},
                "max": {"color": "#00ff00"},
            },
        },
    ]
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=2, cols=3),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps({"version": 1, "conditionalFormats": {"sheet-0": rules}}),
        },
    )

    assert response.status_code == 200
    ws = load_workbook(io.BytesIO(response.content)).active
    exported_rules = [
        rule
        for rules_for_range in ws.conditional_formatting._cf_rules.values()
        for rule in rules_for_range
    ]
    formulas = [rule.formula for rule in exported_rules]
    assert ["AND(A1>1.0,A1<2.0)"] in formulas
    assert ['NOT(ISERROR(FIND("X",B1)))'] in formulas
    assert ['"001"'] in formulas
    color_scales = [rule.colorScale for rule in exported_rules if rule.colorScale is not None]
    numeric_scale = next(
        scale for scale in color_scales if [point.type for point in scale.cfvo] == ["num"] * 3
    )
    production_scale = next(
        scale
        for scale in color_scales
        if [point.type for point in scale.cfvo] == ["min", "percent", "max"]
    )
    assert [point.val for point in numeric_scale.cfvo] == [0.0, 50.0, 100.0]
    assert [point.type for point in production_scale.cfvo] == [
        "min",
        "percent",
        "max",
    ]
    assert production_scale.cfvo[1].val == 50.0


def test_recovery_export_preserves_first_matching_conditional_format(sync_client) -> None:
    rules = [
        {
            "id": "first-match",
            "range": {"top": 0, "left": 0, "bottom": 0, "right": 0},
            "condition": {"kind": "notBlank"},
            "style": {"background": "#ff0000"},
        },
        {
            "id": "second-match",
            "range": {"top": 0, "left": 0, "bottom": 0, "right": 0},
            "condition": {"kind": "notBlank"},
            "style": {"background": "#00ff00"},
        },
    ]
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=1, cols=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps({"version": 1, "conditionalFormats": {"sheet-0": rules}}),
        },
    )

    assert response.status_code == 200
    ws = load_workbook(io.BytesIO(response.content)).active
    exported_rules = [
        rule
        for rules_for_range in ws.conditional_formatting._cf_rules.values()
        for rule in rules_for_range
    ]
    assert [(rule.priority, rule.stopIfTrue) for rule in exported_rules] == [
        (1, True),
        (2, True),
    ]
    assert [rule.dxf.fill.fgColor.rgb for rule in exported_rules] == [
        "FFFF0000",
        "FF00FF00",
    ]


def test_recovery_export_rejects_all_conditional_formats_when_one_is_invalid(
    sync_client,
) -> None:
    rules = [
        {
            "id": "valid-first-match",
            "range": {"top": 0, "left": 0, "bottom": 0, "right": 0},
            "condition": {"kind": "notBlank"},
            "style": {"background": "#ff0000"},
        },
        {
            "id": "invalid-second-match",
            "range": {"top": 0, "left": 0, "bottom": 0, "right": 0},
            "condition": {"kind": "notBlank"},
            "style": {"fontSize": 12},
        },
    ]
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=1, cols=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps({"version": 1, "conditionalFormats": {"sheet-0": rules}}),
        },
    )

    assert response.status_code == 400
    assert response.headers["content-type"].startswith("application/json")


def test_recovery_export_rejects_rule_overlapping_an_earlier_color_scale(
    sync_client,
) -> None:
    rules = [
        {
            "id": "color-scale-first",
            "range": {"top": 0, "left": 0, "bottom": 1, "right": 0},
            "condition": {
                "kind": "colorScale",
                "min": {"color": "#ff0000"},
                "max": {"color": "#00ff00"},
            },
        },
        {
            "id": "overlapping-later-rule",
            "range": {"top": 1, "left": 0, "bottom": 1, "right": 0},
            "condition": {"kind": "notBlank"},
            "style": {"bold": True},
        },
    ]
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=2, cols=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps({"version": 1, "conditionalFormats": {"sheet-0": rules}}),
        },
    )

    assert response.status_code == 400
    assert "color scale" in response.json()["error"]["message"]


def test_recovery_export_rejects_lone_surrogate_in_cell_value(sync_client) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=1, cols=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps(
                {
                    "version": 1,
                    "cells": {"sheet-0!0,0": {"value": "\ud800"}},
                }
            ),
        },
    )

    assert response.status_code == 400
    assert "XML 1.0" in response.json()["error"]["message"]


@pytest.mark.parametrize(
    ("field", "state"),
    [
        (
            "sheet name",
            {
                "version": 1,
                "workbookOps": [
                    {
                        "type": "addSheet",
                        "sheetId": "sheet-new",
                        "name": "Bad\x01Name",
                    }
                ],
            },
        ),
        ("formula", {"version": 1, "cells": {"sheet-0!0,0": {"formula": "=\x01"}}}),
        (
            "number format",
            {"version": 1, "cells": {"sheet-0!0,0": {"numberFormat": "0\x01"}}},
        ),
        (
            "font family",
            {
                "version": 1,
                "cells": {"sheet-0!0,0": {"style": {"fontFamily": "Bad\x01Font"}}},
            },
        ),
        (
            "hyperlink",
            {
                "version": 1,
                "cells": {"sheet-0!0,0": {"style": {"hyperlink": "https://x/\x01"}}},
            },
        ),
        (
            "comment text",
            {
                "version": 1,
                "comments": {"sheet-0!0,0": {"text": "Bad\x01comment"}},
            },
        ),
        (
            "comment author",
            {
                "version": 1,
                "comments": {"sheet-0!0,0": {"text": "Comment", "author": "Bad\x01author"}},
            },
        ),
        (
            "validation value",
            {
                "version": 1,
                "validations": {"sheet-0!0,0": {"type": "list", "values": ["Bad\x01value"]}},
            },
        ),
        (
            "conditional-format text",
            {
                "version": 1,
                "conditionalFormats": {
                    "sheet-0": [
                        {
                            "id": "bad-text",
                            "range": {"top": 0, "left": 0, "bottom": 0, "right": 0},
                            "condition": {
                                "kind": "containsText",
                                "text": "Bad\x01text",
                                "mode": "contains",
                            },
                        }
                    ]
                },
            },
        ),
        (
            "conditional-format operand",
            {
                "version": 1,
                "conditionalFormats": {
                    "sheet-0": [
                        {
                            "id": "bad-operand",
                            "range": {"top": 0, "left": 0, "bottom": 0, "right": 0},
                            "condition": {
                                "kind": "cellValue",
                                "op": "eq",
                                "value": "Bad\x01operand",
                            },
                        }
                    ]
                },
            },
        ),
    ],
)
def test_recovery_export_rejects_invalid_xml_in_every_string_sink(
    sync_client, field, state
) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=1, cols=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"strict_recovery": "true", "edits": json.dumps(state)},
    )

    assert response.status_code == 400, field
    assert "XML 1.0" in response.json()["error"]["message"]


def test_recovery_export_roundtrips_valid_xml_unicode_and_numeric_ceiling(sync_client) -> None:
    cell_text = "Line\tOne\nTwo😀"
    numeric_ceiling = 9.99999999999999e307
    state = {
        "version": 1,
        "workbookOps": [
            {
                "type": "addSheet",
                "sheetId": "sheet-new",
                "name": "恢复😀",
                "afterSheetId": "sheet-0",
            }
        ],
        "cells": {
            "sheet-new!0,0": {"value": cell_text},
            "sheet-new!0,1": {"value": numeric_ceiling},
        },
    }

    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=1, cols=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"strict_recovery": "true", "edits": json.dumps(state)},
    )

    assert response.status_code == 200
    wb = load_workbook(io.BytesIO(response.content), data_only=False)
    assert wb.sheetnames == ["Sheet", "恢复😀"]
    assert wb["恢复😀"]["A1"].value == cell_text
    assert wb["恢复😀"]["B1"].value == numeric_ceiling


def test_recovery_export_preserves_equals_prefixed_value_as_text(sync_client) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=1, cols=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps(
                {"version": 1, "cells": {"sheet-0!0,0": {"value": "=literal text"}}}
            ),
        },
    )

    assert response.status_code == 200
    cell = load_workbook(io.BytesIO(response.content), data_only=False).active["A1"]
    assert cell.value == "=literal text"
    assert cell.data_type == "s"


def test_recovery_export_rejects_validation_overlapping_source_rule(sync_client) -> None:
    wb = Workbook()
    ws = wb.active
    validation = DataValidation(type="list", formula1='"Source"')
    validation.add("A1")
    ws.add_data_validation(validation)
    source = io.BytesIO()
    wb.save(source)

    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                source.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps(
                {
                    "version": 1,
                    "validations": {"sheet-0!0,0": {"type": "list", "values": ["Recovered"]}},
                }
            ),
        },
    )

    assert response.status_code == 400
    assert "existing source validation" in response.json()["error"]["message"]


def test_recovery_export_escapes_search_wildcards_as_literal_text(sync_client) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=1, cols=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "strict_recovery": "true",
            "edits": json.dumps(
                {
                    "version": 1,
                    "conditionalFormats": {
                        "sheet-0": [
                            {
                                "id": "literal-search",
                                "range": {"top": 0, "left": 0, "bottom": 0, "right": 0},
                                "condition": {
                                    "kind": "containsText",
                                    "text": "a~*?b",
                                    "mode": "contains",
                                },
                            }
                        ]
                    },
                }
            ),
        },
    )

    assert response.status_code == 200
    ws = load_workbook(io.BytesIO(response.content), data_only=False).active
    rules = [
        rule
        for rules_for_range in ws.conditional_formatting._cf_rules.values()
        for rule in rules_for_range
    ]
    assert ['NOT(ISERROR(SEARCH("a~~~*~?b",A1)))'] in [rule.formula for rule in rules]


@pytest.mark.parametrize(
    "condition",
    [
        {"kind": "containsText", "text": "*" * 4_090, "mode": "contains"},
        {"kind": "cellValue", "op": "eq", "value": '"' * 4_096},
    ],
)
def test_recovery_export_rejects_generated_conditional_format_formulas_over_limit(
    sync_client, condition
) -> None:
    state = {
        "version": 1,
        "conditionalFormats": {
            "sheet-0": [
                {
                    "id": "too-long",
                    "range": {"top": 0, "left": 0, "bottom": 0, "right": 0},
                    "condition": condition,
                }
            ]
        },
    }
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=1, cols=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"strict_recovery": "true", "edits": json.dumps(state)},
    )

    assert response.status_code == 400
    assert "8192" in response.json()["error"]["message"]


@pytest.mark.parametrize(
    "state",
    [
        {"version": 1, "cells": {"sheet-0!0,0": {"value": 10**4_000}}},
        {
            "version": 1,
            "cells": {"sheet-0!0,0": {"value": 1.7976931348623157e308}},
        },
        {"version": 1, "cells": {"sheet-0!0,0": {"value": 1.2345678901234567}}},
        {"version": 1, "cells": {"sheet-0!0,0": {"value": 5e-324}}},
        {
            "version": 1,
            "conditionalFormats": {
                "sheet-0": [
                    {
                        "id": "huge-operand",
                        "range": {"top": 0, "left": 0, "bottom": 0, "right": 0},
                        "condition": {"kind": "cellValue", "op": "gt", "value": 10**4_000},
                    }
                ]
            },
        },
        {
            "version": 1,
            "conditionalFormats": {
                "sheet-0": [
                    {
                        "id": "rounded-operand",
                        "range": {"top": 0, "left": 0, "bottom": 0, "right": 0},
                        "condition": {
                            "kind": "cellValue",
                            "op": "gt",
                            "value": 1.2345678901234567,
                        },
                    }
                ]
            },
        },
    ],
)
def test_recovery_export_rejects_numeric_values_xlsx_cannot_preserve(sync_client, state) -> None:
    response = sync_client.post(
        "/api/excel/export-edited",
        files={
            "file": (
                "workbook.xlsx",
                _build_xlsx(rows=1, cols=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"strict_recovery": "true", "edits": json.dumps(state)},
    )

    assert response.status_code == 400
